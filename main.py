from tkinter import *
from openpyxl import Workbook, load_workbook
try:
    wb=load_workbook("wb.xlsx")
except:
    wb=Workbook()
sheet=wb.active
import os
def excel():
    global wb,sheet
    if not os.path.exists("wb.xlsx"):
        wb = Workbook()
        ws = wb.active
        ws.title = "Student Data"

        ws.append([
          "Name",
          "Course",
          "Semester",
          "Form NO",
          "Contact Number",
          "Email ID",
          "Address"
        ])

        wb.save("wb.xlsx")
        wb = load_workbook("wb.xlsx")
        sheet = wb.active
def focus0(event):
    name_field.focus_set()
def focus1(event):
    course_field.focus_set()
def focus2(event):
    sem_field.focus_set()
def focus3(event):
    form_no_field.focus_set()
def focus4(event):
    contact_no_field.focus_set()
def focus5(event):
    email_id_field.focus_set()
def focus6(event):
    address_field.focus_set()
def clear():
    name_field.delete(0, END)
    course_field.delete(0, END)
    sem_field.delete(0, END)
    form_no_field.delete(0, END)
    contact_no_field.delete(0, END)
    email_id_field.delete(0, END)
    address_field.delete(0, END)
def insert():
    if (name_field.get()==""and
    course_field.get()==""and
    sem_field.get()==""and
    form_no_field.get()==""and
    contact_no_field.get()==""and
    email_id_field.get()==""and
    address_field.get()==""):
        print("Empty Field")
    else:
        current_row = sheet.max_row
        current_column = sheet.max_column
        sheet.cell(row=current_row + 1, column=1).value = name_field.get()
        sheet.cell(row=current_row + 1, column=2).value = course_field.get()
        sheet.cell(row=current_row + 1, column=3).value = sem_field.get()
        sheet.cell(row=current_row + 1, column=4).value = form_no_field.get()
        sheet.cell(row=current_row + 1, column=5).value = contact_no_field.get()
        sheet.cell(row=current_row + 1, column=6).value = email_id_field.get()
        sheet.cell(row=current_row + 1, column=7).value = address_field.get()
        wb.save("wb.xlsx")
        name_field.focus_set()
        clear()
if __name__ == "__main__":
    root = Tk()
    root.title("Registration Form")
    root.geometry("500x400")
    root.configure(bg="light grey")
    excel()
    heading = Label(root, text="Form", bg= "grey")
    name = Label(root, text= "Name", bg="light grey")
    course = Label(root,text= "Course", bg="light grey")
    sem = Label(root,text= "Semester", bg= "light grey")
    form_no = Label(root, text= "Form N0.", bg= "light grey")
    contact_no = Label(root, text= "Contact Number", bg = "light grey")
    email_id = Label(root, text= "Email-ID",bg="light grey")
    address = Label(root, text= "Address", bg = "light grey")
    name.grid(row=1, column=0)
    course.grid(row=2, column=0)
    sem.grid(row=3, column=0)
    form_no.grid(row=4, column=0)
    contact_no.grid(row=5, column=0)
    email_id.grid(row=6, column=0)
    address.grid(row=7, column=0)
    name_field = Entry(root)
    course_field=Entry(root)
    sem_field=Entry(root)
    form_no_field=Entry(root)
    contact_no_field=Entry(root)
    email_id_field=Entry(root)
    address_field=Entry(root)
    name_field.bind("<Return>",focus0)
    course_field.bind("<Return>",focus1)
    sem_field.bind("<Return>;", focus2)
    form_no_field.bind("<Return>", focus3)
    contact_no_field.bind("<Return>",focus4)
    email_id_field.bind("<Return>", focus5)
    address_field.bind("<Return>", focus6)
    name_field.grid(row=1, column=1, ipadx="100")
    course_field.grid(row=2, column=1, ipadx="100")
    sem_field.grid(row=3, column=1, ipadx="100")
    form_no_field.grid(row=4, column=1, ipadx="100")
    contact_no_field.grid(row=5, column=1, ipadx="100")
    email_id_field.grid(row=6, column=1, ipadx="100")
    address_field.grid(row=7, column=1, ipadx="100")
    excel()
    submit = Button(root, text="Submit", fg="Black",bg="blue", command=insert)
    submit.grid(row=8, column=1)
    root.mainloop()

